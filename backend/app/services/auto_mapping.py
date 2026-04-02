"""
Auto-mapping service — WP-6b.

Analyzes uploaded files and proposes column-to-property mappings
without requiring user-provided configuration.

Pure functions, no state. Takes file bytes and optional hints,
returns a proposed mapping with confidence scores.

Architecture (Section 3.3 of WP-6b spec):
  1. detect_header_row()    — find the header row in the file
  2. detect_target_type()   — score registered types against headers
  3. build_property_mapping() — three-layer matching per column
  4. detect_identifier_column() — find the identifier column
  5. propose_mapping()      — orchestrator returning ProposedMapping
"""

import io
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from typing import Any

import openpyxl

from app.core.column_aliases import (
    HEADER_KEYWORDS,
    IDENTIFIER_ALIASES,
    clean_column_name,
    get_aliases_for_type,
    is_ignored_column,
)
from app.core.type_config import (
    TypeConfig,
    build_label_map,
    get_dimension_properties,
    get_importable_types,
    get_type_config,
)
from app.schemas.imports import ImportMappingConfig


def _build_label_map_from_config(tc: TypeConfig) -> dict[str, str]:
    """Build label map directly from a TypeConfig (for dynamic types not in ITEM_TYPES)."""
    label_map: dict[str, str] = {}
    for prop in tc.properties:
        label_map[prop.label.lower()] = prop.name
        label_map[prop.name] = prop.name
        if prop.aliases:
            for alias in prop.aliases:
                label_map[alias.lower()] = prop.name
    return label_map


# ─── Data Classes ────────────────────────────────────────────────


@dataclass
class ColumnProposal:
    """Proposed mapping for a single column."""

    column_name: str  # Raw column header text
    cleaned_name: str  # Cleaned/normalized column name
    proposed_property: str | None = None  # Proposed Cadence property name
    confidence: float = 0.0  # 0.0–1.0
    match_method: str = "none"  # "exact_label", "normalized_label", "alias", "fuzzy", "identifier", "ignored", "none"
    alternatives: list[str] = field(default_factory=list)


@dataclass
class ProposedMapping:
    """Complete auto-mapping proposal for a file."""

    header_row: int  # 1-indexed
    header_row_confidence: float
    target_item_type: str
    type_confidence: float
    identifier_column: str  # Raw column header
    identifier_confidence: float
    columns: list[ColumnProposal] = field(default_factory=list)
    unmatched_columns: list[str] = field(default_factory=list)
    proposed_config: ImportMappingConfig | None = None
    overall_confidence: float = 0.0
    needs_user_review: bool = True


# ─── Header Row Detection ────────────────────────────────────────


def detect_header_row(
    file_bytes: bytes,
    file_type: str = "excel",
    max_rows_to_check: int = 20,
) -> tuple[int, float]:
    """
    Detect the most likely header row in a file.

    Ported from Delta Project ExcelProcessor._detect_header_row,
    enhanced with type-registry awareness.

    Scoring:
      - Text-to-numeric ratio (headers are text, not numbers)
      - AEC keyword matches (number, type, width, height, etc.)
      - Identifier field matches (mark, number, id, etc.)
      - Row depth penalty (headers are usually near the top)
      - Non-empty cell bonus (headers have many filled cells)

    Returns: (row_index_1based, confidence 0.0–1.0)
    """
    if file_type == "csv":
        return _detect_header_row_csv(file_bytes, max_rows_to_check)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Read sample rows
    sample_rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows_to_check:
            break
        sample_rows.append(list(row))

    wb.close()

    if not sample_rows:
        return (1, 0.0)

    best_row = 0
    best_score = -999.0

    for idx, row in enumerate(sample_rows):
        row_values = [
            str(v).lower().strip() for v in row if v is not None and str(v).strip()
        ]

        if not row_values:
            continue

        score = 0.0

        # Text-to-numeric ratio: headers are mostly text
        text_count = sum(
            1 for v in row_values if not v.replace(".", "").replace("-", "").isdigit()
        )
        text_ratio = text_count / max(1, len(row_values))
        score += text_ratio * 5

        # AEC keyword matches
        keyword_matches = sum(
            1 for v in row_values if any(kw in v for kw in HEADER_KEYWORDS)
        )
        score += keyword_matches * 2

        # Identifier field matches (strong signal)
        cleaned_values = [clean_column_name(v) for v in row_values]
        id_field_matches = sum(1 for v in cleaned_values if v in IDENTIFIER_ALIASES)
        score += id_field_matches * 3

        # Non-empty cell count bonus (headers tend to have many filled cells)
        non_empty_count = len(row_values)
        total_cells = len(row)
        if total_cells > 0:
            fill_ratio = non_empty_count / total_cells
            score += fill_ratio * 2

        # Row depth penalty (earlier rows more likely to be headers)
        score -= idx * 0.5

        if score > best_score:
            best_score = score
            best_row = idx

    # Normalize confidence: map score range to 0.0–1.0
    # Typical header rows score 10–20; non-headers score < 5
    confidence = min(1.0, max(0.0, best_score / 15.0))

    return (best_row + 1, confidence)  # Convert to 1-indexed


def _detect_header_row_csv(
    file_bytes: bytes,
    max_rows_to_check: int = 20,
) -> tuple[int, float]:
    """Detect header row in a CSV file. Same algorithm, different parser."""
    import csv

    text_content = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text_content))

    sample_rows: list[list[str]] = []
    for i, row in enumerate(reader):
        if i >= max_rows_to_check:
            break
        sample_rows.append(row)

    if not sample_rows:
        return (1, 0.0)

    best_row = 0
    best_score = -999.0

    for idx, row in enumerate(sample_rows):
        row_values = [v.lower().strip() for v in row if v.strip()]

        if not row_values:
            continue

        score = 0.0
        text_count = sum(
            1 for v in row_values if not v.replace(".", "").replace("-", "").isdigit()
        )
        text_ratio = text_count / max(1, len(row_values))
        score += text_ratio * 5

        keyword_matches = sum(
            1 for v in row_values if any(kw in v for kw in HEADER_KEYWORDS)
        )
        score += keyword_matches * 2

        cleaned_values = [clean_column_name(v) for v in row_values]
        id_field_matches = sum(1 for v in cleaned_values if v in IDENTIFIER_ALIASES)
        score += id_field_matches * 3

        non_empty_count = len(row_values)
        total_cells = len(row)
        if total_cells > 0:
            fill_ratio = non_empty_count / total_cells
            score += fill_ratio * 2

        score -= idx * 0.5

        if score > best_score:
            best_score = score
            best_row = idx

    confidence = min(1.0, max(0.0, best_score / 15.0))
    return (best_row + 1, confidence)


# ─── Target Type Detection ───────────────────────────────────────


def detect_target_type(
    headers: list[str],
    user_aliases: dict[str, str] | None = None,
    importable_types: list[TypeConfig] | None = None,
) -> tuple[str, float, dict[str, int]]:
    """
    Determine which registered item type best matches the column headers.

    Scores each importable type by counting how many of its PropertyDef
    labels (or aliases) appear in the headers.

    Args:
        headers: List of raw column header strings
        user_aliases: Optional project-level user corrections
        importable_types: Optional list of TypeConfig objects to match against.
            When provided, uses these instead of the global ITEM_TYPES registry.

    Returns:
        (type_name, confidence, match_counts_per_type)
    """
    cleaned_headers = [clean_column_name(h) for h in headers]
    # Also keep lowercased raw headers for label matching
    lower_headers = [h.lower().strip() for h in headers]

    types_to_check = importable_types if importable_types is not None else get_importable_types()
    if not types_to_check:
        return ("", 0.0, {})

    match_counts: dict[str, int] = {}

    for tc in types_to_check:
        label_map = _build_label_map_from_config(tc)
        domain_aliases = get_aliases_for_type(tc.name)

        matches = 0
        len(tc.properties)

        for cleaned, lower in zip(cleaned_headers, lower_headers):
            if is_ignored_column(cleaned):
                continue
            if cleaned in IDENTIFIER_ALIASES:
                matches += 1  # Identifier match counts toward type
                continue

            # Check label map (PropertyDef labels + aliases)
            if cleaned in label_map or lower in label_map:
                matches += 1
                continue

            # Check domain aliases
            if cleaned in domain_aliases:
                matches += 1
                continue

            # Check user aliases
            if user_aliases and cleaned in user_aliases:
                matches += 1
                continue

        match_counts[tc.name] = matches

    if not match_counts:
        return ("", 0.0, match_counts)

    # Best type is the one with most matches
    best_type = max(match_counts, key=match_counts.get)  # type: ignore[arg-type]
    best_count = match_counts[best_type]

    # Confidence: ratio of matched columns to total non-ignored columns
    non_ignored = sum(1 for c in cleaned_headers if not is_ignored_column(c))
    confidence = best_count / max(1, non_ignored)

    return (best_type, confidence, match_counts)


# ─── Column Property Mapping ─────────────────────────────────────


def build_property_mapping(
    headers: list[str],
    target_type: str,
    user_aliases: dict[str, str] | None = None,
    importable_types: list[TypeConfig] | None = None,
) -> list[ColumnProposal]:
    """
    For each column header, propose a property mapping on the target type.

    Three-layer matching (Section 3.2):
      1. User corrections (highest priority)
      2. Type-derived labels (PropertyDef labels + aliases)
      3. AEC preset aliases (from column_aliases.py)

    If none of those match, attempts fuzzy string similarity as a
    last resort (confidence < 0.5).

    Args:
        headers: Raw column header strings
        target_type: The target item type name
        user_aliases: Optional project-level user corrections
        importable_types: Optional list of TypeConfig objects. When provided,
            looks up the target type from this list instead of ITEM_TYPES.

    Returns:
        List of ColumnProposal, one per header
    """
    if importable_types is not None:
        tc = next((t for t in importable_types if t.name == target_type), None)
        label_map = _build_label_map_from_config(tc) if tc else {}
    else:
        label_map = build_label_map(target_type)
        tc = get_type_config(target_type)
    domain_aliases = get_aliases_for_type(target_type)
    property_names = [p.name for p in tc.properties] if tc else []

    proposals: list[ColumnProposal] = []
    used_properties: set[str] = set()  # Prevent duplicate mappings

    for raw_header in headers:
        cleaned = clean_column_name(raw_header)
        lower = raw_header.lower().strip()

        proposal = ColumnProposal(
            column_name=raw_header,
            cleaned_name=cleaned,
        )

        # Skip ignored columns
        if is_ignored_column(cleaned):
            proposal.match_method = "ignored"
            proposal.confidence = 1.0
            proposals.append(proposal)
            continue

        # Skip identifier columns (handled separately)
        if cleaned in IDENTIFIER_ALIASES:
            proposal.match_method = "identifier"
            proposal.proposed_property = "__identifier__"
            proposal.confidence = 1.0
            proposals.append(proposal)
            continue

        # ── Layer 1: User corrections (highest priority) ─────────
        if user_aliases and cleaned in user_aliases:
            prop = user_aliases[cleaned]
            if prop not in used_properties:
                proposal.proposed_property = prop
                proposal.confidence = 1.0
                proposal.match_method = "user_correction"
                used_properties.add(prop)
                proposals.append(proposal)
                continue

        # ── Layer 2: Type-derived label match ────────────────────
        # Check exact label match (lowercased)
        matched_prop = label_map.get(lower) or label_map.get(cleaned)
        if matched_prop and matched_prop not in used_properties:
            # Determine if it was an exact label or an alias from the PropertyDef
            exact_label = False
            if tc:
                for p in tc.properties:
                    if p.label.lower() == lower or p.name == cleaned:
                        exact_label = True
                        break

            proposal.proposed_property = matched_prop
            proposal.confidence = 1.0 if exact_label else 0.85
            proposal.match_method = "exact_label" if exact_label else "normalized_label"
            used_properties.add(matched_prop)
            proposals.append(proposal)
            continue

        # ── Layer 3: AEC preset aliases ──────────────────────────
        alias_prop = domain_aliases.get(cleaned)
        if alias_prop and alias_prop not in used_properties:
            proposal.proposed_property = alias_prop
            proposal.confidence = 0.75
            proposal.match_method = "alias"
            used_properties.add(alias_prop)
            proposals.append(proposal)
            continue

        # ── Layer 4: Fuzzy matching (last resort) ────────────────
        best_fuzzy_score = 0.0
        best_fuzzy_prop = None
        alternatives: list[str] = []

        for prop_name in property_names:
            if prop_name in used_properties:
                continue
            # Compare cleaned header against property name and label
            score_name = SequenceMatcher(None, cleaned, prop_name).ratio()
            prop_def_label = ""
            if tc:
                for p in tc.properties:
                    if p.name == prop_name:
                        prop_def_label = p.label.lower()
                        break
            score_label = (
                SequenceMatcher(None, lower, prop_def_label).ratio()
                if prop_def_label
                else 0.0
            )
            best_score = max(score_name, score_label)

            if best_score > 0.6:
                alternatives.append(prop_name)
                if best_score > best_fuzzy_score:
                    best_fuzzy_score = best_score
                    best_fuzzy_prop = prop_name

        if best_fuzzy_prop and best_fuzzy_score > 0.6:
            proposal.proposed_property = best_fuzzy_prop
            proposal.confidence = best_fuzzy_score * 0.7  # Downweight fuzzy matches
            proposal.match_method = "fuzzy"
            proposal.alternatives = [a for a in alternatives if a != best_fuzzy_prop]
            used_properties.add(best_fuzzy_prop)
        else:
            proposal.alternatives = alternatives

        proposals.append(proposal)

    return proposals


# ─── Identifier Column Detection ─────────────────────────────────


def detect_identifier_column(
    headers: list[str],
    user_aliases: dict[str, str] | None = None,
) -> tuple[str, float]:
    """
    Detect which column contains item identifiers.

    Checks headers against the known identifier field set,
    with fallback heuristics.

    Returns: (raw_column_name, confidence 0.0–1.0)
    """
    # First pass: exact match against identifier aliases
    for raw_header in headers:
        cleaned = clean_column_name(raw_header)
        if cleaned in IDENTIFIER_ALIASES:
            return (raw_header, 1.0)

    # Second pass: fuzzy match against identifier aliases
    for raw_header in headers:
        cleaned = clean_column_name(raw_header)
        for id_alias in IDENTIFIER_ALIASES:
            score = SequenceMatcher(None, cleaned, id_alias).ratio()
            if score > 0.8:
                return (raw_header, score * 0.8)

    # Fallback: first text column
    if headers:
        return (headers[0], 0.3)

    return ("", 0.0)


# ─── File Header Extraction ──────────────────────────────────────


def extract_headers(
    file_bytes: bytes,
    file_type: str = "excel",
    header_row: int = 1,
) -> list[str]:
    """
    Extract column headers from a file at the specified row.

    Args:
        file_bytes: Raw file content
        file_type: "excel" or "csv"
        header_row: 1-indexed row number

    Returns:
        List of header strings
    """
    if file_type == "csv":
        return _extract_headers_csv(file_bytes, header_row)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows_iter, None)
    header_values = next(rows_iter, None)
    wb.close()

    if not header_values:
        return []

    return [str(h).strip() if h is not None else "" for h in header_values]


def _extract_headers_csv(file_bytes: bytes, header_row: int) -> list[str]:
    """Extract headers from a CSV file."""
    import csv

    text_content = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text_content))

    for _ in range(header_row - 1):
        next(reader, None)
    header_values = next(reader, None)

    if not header_values:
        return []

    return [h.strip() for h in header_values]


# ─── Orchestrator ─────────────────────────────────────────────────


def propose_mapping(
    file_bytes: bytes,
    file_type: str = "excel",
    user_aliases: dict[str, str] | None = None,
    project_id: str | None = None,
    importable_types: list[TypeConfig] | None = None,
) -> ProposedMapping:
    """
    Analyze a file and propose a complete column-to-property mapping.

    Orchestrates all detection functions:
      1. Detect header row
      2. Extract headers
      3. Detect target type
      4. Detect identifier column
      5. Build property mapping for each column
      6. Assemble ProposedMapping with ImportMappingConfig

    Args:
        file_bytes: Raw file content
        file_type: "excel" or "csv"
        user_aliases: Optional project-level user corrections
        project_id: Optional project ID for scoped alias lookup

    Returns:
        ProposedMapping with proposed config and confidence scores
    """
    # Step 1: Detect header row
    header_row, header_confidence = detect_header_row(file_bytes, file_type)

    # Step 2: Extract headers
    headers = extract_headers(file_bytes, file_type, header_row)
    if not headers:
        return ProposedMapping(
            header_row=header_row,
            header_row_confidence=header_confidence,
            target_item_type="",
            type_confidence=0.0,
            identifier_column="",
            identifier_confidence=0.0,
            overall_confidence=0.0,
            needs_user_review=True,
        )

    # Step 3: Detect target type
    target_type, type_confidence, _ = detect_target_type(
        headers, user_aliases, importable_types=importable_types,
    )

    # Step 4: Detect identifier column
    identifier_col, identifier_confidence = detect_identifier_column(
        headers, user_aliases
    )

    # Step 5: Build property mapping
    if target_type:
        column_proposals = build_property_mapping(
            headers, target_type, user_aliases, importable_types=importable_types,
        )
    else:
        column_proposals = [
            ColumnProposal(column_name=h, cleaned_name=clean_column_name(h))
            for h in headers
        ]

    # Step 6: Assemble results
    unmatched = [
        cp.column_name
        for cp in column_proposals
        if cp.match_method == "none" and cp.proposed_property is None
    ]

    # Build the ImportMappingConfig
    property_mapping: dict[str, str] = {}
    normalizations: dict[str, str] = {}

    # Get dimension properties for normalization assignment
    get_dimension_properties(target_type) if target_type else set()
    if importable_types is not None:
        tc = next((t for t in importable_types if t.name == target_type), None)
    else:
        tc = get_type_config(target_type)

    for cp in column_proposals:
        if (
            cp.proposed_property
            and cp.proposed_property != "__identifier__"
            and cp.match_method != "ignored"
        ):
            property_mapping[cp.column_name] = cp.proposed_property

            # Auto-assign normalization based on PropertyDef
            if tc:
                for prop_def in tc.properties:
                    if prop_def.name == cp.proposed_property:
                        if prop_def.normalization:
                            normalizations[cp.proposed_property] = (
                                prop_def.normalization
                            )
                        elif prop_def.unit:
                            normalizations[cp.proposed_property] = "dimension"
                        break

    proposed_config = (
        ImportMappingConfig(
            file_type=file_type,
            identifier_column=identifier_col,
            target_item_type=target_type,
            header_row=header_row,
            property_mapping=property_mapping,
            normalizations=normalizations,
        )
        if target_type and identifier_col
        else None
    )

    # Calculate overall confidence
    mapped_confidences = [
        cp.confidence
        for cp in column_proposals
        if cp.match_method not in ("ignored", "identifier", "none")
    ]
    if mapped_confidences:
        avg_confidence = sum(mapped_confidences) / len(mapped_confidences)
    else:
        avg_confidence = 0.0

    overall = (
        header_confidence * 0.15
        + type_confidence * 0.25
        + identifier_confidence * 0.15
        + avg_confidence * 0.45
    )

    # Needs review if any mapped column has low confidence, or if there are unmatched columns
    needs_review = (
        any(
            cp.confidence < 0.8
            for cp in column_proposals
            if cp.match_method not in ("ignored",)
        )
        or len(unmatched) > 0
        or overall < 0.8
    )

    return ProposedMapping(
        header_row=header_row,
        header_row_confidence=header_confidence,
        target_item_type=target_type,
        type_confidence=type_confidence,
        identifier_column=identifier_col,
        identifier_confidence=identifier_confidence,
        columns=column_proposals,
        unmatched_columns=unmatched,
        proposed_config=proposed_config,
        overall_confidence=overall,
        needs_user_review=needs_review,
    )
