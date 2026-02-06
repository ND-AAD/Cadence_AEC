"""
Tests for WP-9: Change Detection on Import.

Covers:
  - Change detection runs AFTER snapshot creation
  - Compares new values against prior source values
  - Uses normalized comparison (values_match)
  - One change item per (source, affected_item) containing all property changes
  - Change item is self-sourced
  - Correct identifier format for change items
  - Correct connections (4 total)
  - Summary counts accurate
  - No false changes on unchanged properties
  - First import (no prior context) → 0 change items
  - Handles multiple property changes on same item
"""

import json
import uuid

import pytest
import pytest_asyncio
from httpx import AsyncClient
from sqlalchemy import func, select

from app.models.core import Connection, Item, Snapshot
from app.services.import_service import _find_prior_context
from tests.fixtures.excel_factory import (
    STANDARD_DOOR_MAPPING,
    make_door_schedule_excel,
    make_updated_door_schedule_excel,
)


# ─── Helpers ──────────────────────────────────────────────────


@pytest_asyncio.fixture
async def project_setup(make_item, make_connection):
    """
    Create a minimal project with source and milestones.

    Returns dict with: project, schedule, dd_milestone, cd_milestone, bd_milestone
    """
    project = await make_item("project", "Project Alpha")
    schedule = await make_item(
        "schedule",
        "Finish Schedule",
        {"name": "Finish Schedule", "discipline": "Architectural"},
    )
    dd_milestone = await make_item(
        "milestone",
        "DD",
        {"name": "Design Development", "ordinal": 100},
    )
    cd_milestone = await make_item(
        "milestone",
        "CD",
        {"name": "Construction Documents", "ordinal": 200},
    )
    bd_milestone = await make_item(
        "milestone",
        "BD",
        {"name": "Building Design", "ordinal": 300},
    )
    # Wire up connections
    await make_connection(project, schedule)
    await make_connection(project, dd_milestone)
    await make_connection(project, cd_milestone)
    await make_connection(project, bd_milestone)
    return {
        "project": project,
        "schedule": schedule,
        "dd_milestone": dd_milestone,
        "cd_milestone": cd_milestone,
        "bd_milestone": bd_milestone,
    }


# ─── Tests: Basic Change Detection ─────────────────────────────


@pytest.mark.asyncio
async def test_import_first_time_no_prior_context(client: AsyncClient, project_setup):
    """
    Import at DD (no prior context) → 0 change items.

    This is the first import, so there are no prior snapshots to compare against.
    """
    setup = project_setup
    file_bytes = make_door_schedule_excel(5)

    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("door_schedule.xlsx", file_bytes, "application/octet-stream")},
    )
    assert resp.status_code == 201
    result = resp.json()

    assert result["summary"]["source_changes"] == 0
    assert result["summary"]["affected_items"] == 0
    assert len(result["change_items"]) == 0


@pytest.mark.asyncio
async def test_import_detects_changes_on_reimport(client: AsyncClient, project_setup):
    """
    Import at DD, then import updated file at CD → change items for changed properties.

    First import: Door 001-005 with standard finish (paint, stain, etc.)
    Second import: Door 001-005 with updated finish (all stain)
    Changes detected on finish property.
    """
    setup = project_setup
    original_file = make_door_schedule_excel(5)
    updated_file = make_updated_door_schedule_excel(5, changed_indices=list(range(1, 6)), changed_finish="lacquer")

    # First import at DD
    resp1 = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", original_file, "application/octet-stream")},
    )
    assert resp1.status_code == 201

    # Second import at CD with updated file
    resp2 = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )
    assert resp2.status_code == 201
    result2 = resp2.json()

    # All 5 doors should have changes (finish changed)
    assert result2["summary"]["affected_items"] == 5
    assert result2["summary"]["source_changes"] == 5  # 5 doors × 1 property = 5 changes
    assert len(result2["change_items"]) == 5


@pytest.mark.asyncio
async def test_import_no_false_changes_on_unchanged_properties(
    client: AsyncClient, project_setup
):
    """
    Import same file twice at different milestones → 0 changes.

    Even though time_context changes (DD → CD), the property values are identical,
    so no changes should be detected.
    """
    setup = project_setup
    file_bytes = make_door_schedule_excel(5)

    # First import at DD
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )

    # Second import same file at CD
    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )
    assert resp.status_code == 201
    result = resp.json()

    assert result["summary"]["source_changes"] == 0
    assert result["summary"]["affected_items"] == 0
    assert len(result["change_items"]) == 0


# ─── Tests: Change Item Structure ─────────────────────────────


@pytest.mark.asyncio
async def test_change_item_has_correct_identifier(
    client: AsyncClient, project_setup, db_session
):
    """
    Change item identifier format: "{source} / {affected_item} / {from}→{to}"
    """
    from sqlalchemy import select

    setup = project_setup
    original_file = make_door_schedule_excel(1)
    updated_file = make_updated_door_schedule_excel(1, changed_indices=[1], changed_finish="lacquer")

    # First import at DD
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", original_file, "application/octet-stream")},
    )

    # Second import at CD
    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )
    assert resp.status_code == 201

    # Fetch the change item
    result = await db_session.execute(
        select(Item).where(Item.item_type == "change")
    )
    change_item = result.scalar_one_or_none()
    assert change_item is not None

    # Check identifier format
    expected_part = "Finish Schedule / Door 001 / DD→CD"
    assert expected_part in change_item.identifier


@pytest.mark.asyncio
async def test_change_item_has_self_sourced_snapshot(
    client: AsyncClient, project_setup, db_session
):
    """
    Change item has self-sourced snapshot: source_id = item_id (change item itself)
    """
    from sqlalchemy import and_, select

    setup = project_setup
    original_file = make_door_schedule_excel(1)
    updated_file = make_updated_door_schedule_excel(1, changed_indices=[1], changed_finish="lacquer")

    # First import at DD
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", original_file, "application/octet-stream")},
    )

    # Second import at CD
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )

    # Fetch change item and its snapshot
    change_result = await db_session.execute(
        select(Item).where(Item.item_type == "change")
    )
    change_item = change_result.scalar_one()

    snap_result = await db_session.execute(
        select(Snapshot).where(
            and_(
                Snapshot.item_id == change_item.id,
                Snapshot.source_id == change_item.id,  # Self-sourced
            )
        )
    )
    snap = snap_result.scalar_one()
    assert snap is not None
    assert snap.context_id == setup["cd_milestone"].id


@pytest.mark.asyncio
async def test_change_item_snapshot_has_correct_properties(
    client: AsyncClient, project_setup, db_session
):
    """
    Change item snapshot has: status, changes (dict), from_context, to_context, source, affected_item
    """
    from sqlalchemy import and_, select

    setup = project_setup
    original_file = make_door_schedule_excel(1)
    updated_file = make_updated_door_schedule_excel(1, changed_indices=[1], changed_finish="lacquer")

    # First import at DD
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", original_file, "application/octet-stream")},
    )

    # Second import at CD
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )

    # Fetch change item snapshot
    change_result = await db_session.execute(
        select(Item).where(Item.item_type == "change")
    )
    change_item = change_result.scalar_one()

    snap_result = await db_session.execute(
        select(Snapshot).where(
            Snapshot.item_id == change_item.id,
            Snapshot.source_id == change_item.id,
        )
    )
    snap = snap_result.scalar_one()

    # Check properties
    assert snap.properties["status"] == "DETECTED"
    assert "changes" in snap.properties
    assert isinstance(snap.properties["changes"], dict)
    assert "finish" in snap.properties["changes"]  # The changed property
    assert "from_context" in snap.properties
    assert "to_context" in snap.properties
    assert "source" in snap.properties
    assert "affected_item" in snap.properties


@pytest.mark.asyncio
async def test_change_item_has_four_connections(
    client: AsyncClient, project_setup, db_session
):
    """
    Change item has 4 connections:
      - change → source
      - change → to_context
      - change → from_context
      - change → affected_item
    """
    from sqlalchemy import and_, select

    setup = project_setup
    original_file = make_door_schedule_excel(1)
    updated_file = make_updated_door_schedule_excel(1, changed_indices=[1], changed_finish="lacquer")

    # First import at DD
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", original_file, "application/octet-stream")},
    )

    # Second import at CD
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )

    # Fetch change item
    change_result = await db_session.execute(
        select(Item).where(Item.item_type == "change")
    )
    change_item = change_result.scalar_one()

    # Get all outgoing connections from change item
    conns_result = await db_session.execute(
        select(Connection).where(
            Connection.source_item_id == change_item.id
        )
    )
    conns = conns_result.scalars().all()

    # Should have exactly 4 connections
    assert len(conns) == 4

    # Check target item types
    target_ids = {conn.target_item_id for conn in conns}
    assert setup["schedule"].id in target_ids  # source
    assert setup["cd_milestone"].id in target_ids  # to_context
    assert setup["dd_milestone"].id in target_ids  # from_context
    # affected item is a door (Door 001)
    door_result = await db_session.execute(
        select(Item).where(
            and_(Item.item_type == "door", Item.identifier == "Door 001")
        )
    )
    door = door_result.scalar_one()
    assert door.id in target_ids


# ─── Tests: Normalized Comparison ─────────────────────────────


@pytest.mark.asyncio
async def test_normalized_comparison_no_false_change_on_case_difference(
    client: AsyncClient, project_setup, db_session, make_item
):
    """
    Values that differ only in case should NOT trigger a change.
    Uses values_match from normalization.py
    """
    from sqlalchemy import and_, select

    setup = project_setup

    # Create a door manually with lowercase finish
    door = await make_item("door", "Door 001")

    # First import at DD (finish = "paint")
    file_bytes = make_door_schedule_excel(1)
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )

    # Create a custom updated file with "STAIN" (uppercase, same as original "stain")
    # Door 001 original values: finish="stain", material="hollow metal",
    # hardware_set="HW-2", fire_rating="20 min" — must match all except case
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "DOOR NO.", "WIDTH", "HEIGHT", "FINISH", "MATERIAL", "HARDWARE SET", "FIRE RATING"
    ])
    ws.append([
        "Door 001", "3'-0\"", "7'-0\"", "STAIN", "hollow metal", "HW-2", "20 min"
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    updated_file = buf.getvalue()

    # Second import at CD
    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )
    assert resp.status_code == 201
    result = resp.json()

    # Should have NO changes because "paint" == "PAINT" after normalization
    assert result["summary"]["source_changes"] == 0
    assert len(result["change_items"]) == 0


# ─── Tests: Multiple Property Changes ──────────────────────────


@pytest.mark.asyncio
async def test_multiple_properties_changed_single_change_item(
    client: AsyncClient, project_setup, db_session
):
    """
    If multiple properties change on the same item, create ONE change item
    with all property changes in its snapshot.
    """
    from sqlalchemy import select

    setup = project_setup

    # First import at DD
    file_bytes = make_door_schedule_excel(1)
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )

    # Create a custom updated file with changed finish AND material
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "DOOR NO.", "WIDTH", "HEIGHT", "FINISH", "MATERIAL", "HARDWARE SET", "FIRE RATING"
    ])
    # Door 001 originals: finish="stain", material="hollow metal",
    # hardware_set="HW-2", fire_rating="20 min". Change only finish + material.
    ws.append([
        "Door 001", "3'-0\"", "7'-0\"", "lacquer", "aluminum", "HW-2", "20 min"
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    updated_file = buf.getvalue()

    # Second import at CD
    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )
    assert resp.status_code == 201

    # Fetch change item
    change_result = await db_session.execute(
        select(Item).where(Item.item_type == "change")
    )
    change_items = change_result.scalars().all()

    # Should have exactly 1 change item (not one per property)
    assert len(change_items) == 1
    change_item = change_items[0]

    # Its snapshot should contain both changes
    from sqlalchemy import and_
    snap_result = await db_session.execute(
        select(Snapshot).where(
            and_(
                Snapshot.item_id == change_item.id,
                Snapshot.source_id == change_item.id,
            )
        )
    )
    snap = snap_result.scalar_one()

    changes = snap.properties["changes"]
    assert "finish" in changes
    assert "material" in changes
    assert len(changes) == 2


# ─── Tests: Summary Counts ─────────────────────────────────────


@pytest.mark.asyncio
async def test_summary_source_changes_count(client: AsyncClient, project_setup):
    """
    Summary.source_changes counts total property changes across all items.
    """
    setup = project_setup

    # First import at DD
    file_bytes = make_door_schedule_excel(3)
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )

    # Second import with updated file where all 3 doors change finish
    updated_file = make_updated_door_schedule_excel(3, changed_indices=[1, 2, 3], changed_finish="lacquer")
    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )
    assert resp.status_code == 201
    result = resp.json()

    # 3 doors × 1 changed property = 3 changes
    assert result["summary"]["source_changes"] == 3


@pytest.mark.asyncio
async def test_summary_affected_items_count(client: AsyncClient, project_setup):
    """
    Summary.affected_items counts unique items that had at least one change.
    """
    setup = project_setup

    # First import at DD
    file_bytes = make_door_schedule_excel(5)
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )

    # Second import with only first 3 doors changed
    updated_file = make_updated_door_schedule_excel(5, changed_indices=[1, 2, 3], changed_finish="lacquer")
    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", updated_file, "application/octet-stream")},
    )
    assert resp.status_code == 201
    result = resp.json()

    # 3 doors affected
    assert result["summary"]["affected_items"] == 3


# ─── Tests: _find_prior_context Helper ────────────────────────


@pytest.mark.asyncio
async def test_find_prior_context_returns_max_ordinal_less_than_current(
    db_session, make_item
):
    """
    _find_prior_context returns the milestone with highest ordinal
    that is less than the current context's ordinal.
    """
    # Create source and milestones
    source = await make_item("schedule", "Schedule")
    m100 = await make_item("milestone", "DD", {"ordinal": 100})
    m200 = await make_item("milestone", "CD", {"ordinal": 200})
    m300 = await make_item("milestone", "BD", {"ordinal": 300})

    # Create snapshots at m100 and m200
    from app.models.core import Snapshot

    snap1 = Snapshot(
        item_id=source.id,
        context_id=m100.id,
        source_id=source.id,
        properties={},
    )
    snap2 = Snapshot(
        item_id=source.id,
        context_id=m200.id,
        source_id=source.id,
        properties={},
    )
    db_session.add(snap1)
    db_session.add(snap2)
    await db_session.flush()

    # Query for prior context from m300 perspective
    prior = await _find_prior_context(db_session, source.id, m300, {})
    assert prior is not None
    assert prior.id == m200.id


@pytest.mark.asyncio
async def test_find_prior_context_returns_none_for_first_import(
    db_session, make_item
):
    """
    _find_prior_context returns None if there are no prior snapshots.
    """
    # Create source and milestone
    source = await make_item("schedule", "Schedule")
    m100 = await make_item("milestone", "DD", {"ordinal": 100})

    # No snapshots created yet
    prior = await _find_prior_context(db_session, source.id, m100, {})
    assert prior is None


@pytest.mark.asyncio
async def test_find_prior_context_skips_future_contexts(
    db_session, make_item
):
    """
    _find_prior_context ignores milestones with ordinal >= current.
    """
    # Create source and milestones
    source = await make_item("schedule", "Schedule")
    m100 = await make_item("milestone", "DD", {"ordinal": 100})
    m200 = await make_item("milestone", "CD", {"ordinal": 200})
    m300 = await make_item("milestone", "BD", {"ordinal": 300})

    # Create snapshots at m100, m200, m300
    from app.models.core import Snapshot

    for ctx in [m100, m200, m300]:
        snap = Snapshot(
            item_id=source.id,
            context_id=ctx.id,
            source_id=source.id,
            properties={},
        )
        db_session.add(snap)
    await db_session.flush()

    # Query for prior context from m200 perspective
    # Should skip m300 (future) and return m100 (past)
    prior = await _find_prior_context(db_session, source.id, m200, {})
    assert prior is not None
    assert prior.id == m100.id


# ─── Tests: Edge Cases ───────────────────────────────────────


@pytest.mark.asyncio
async def test_import_with_no_prior_snapshots_on_item(
    client: AsyncClient, project_setup, make_item
):
    """
    If an item exists but has no prior snapshot from the source,
    no change should be detected for that item.
    """
    setup = project_setup

    # Create a pre-existing door (no prior snapshot)
    await make_item("door", "Door 999")

    # First import at DD
    file_bytes = make_door_schedule_excel(1)
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )

    # Second import same file at CD
    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )
    assert resp.status_code == 201
    result = resp.json()

    # Door 001 should have no changes (values are identical)
    # Door 999 is not in the import, so no change item for it
    assert result["summary"]["source_changes"] == 0


@pytest.mark.asyncio
async def test_change_items_only_from_prior_source(
    client: AsyncClient, project_setup, db_session, make_item, make_connection
):
    """
    Change detection only compares against prior snapshots from the SAME source.
    If another source has snapshots, they are ignored.
    """
    from sqlalchemy import and_, select

    setup = project_setup

    # Create a second schedule (different source)
    other_schedule = await make_item(
        "schedule",
        "Other Schedule",
        {"name": "Other Schedule", "discipline": "Structural"},
    )
    await make_connection(setup["project"], other_schedule)

    # First import from schedule1 at DD
    file_bytes = make_door_schedule_excel(1)
    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(setup["schedule"].id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", file_bytes, "application/octet-stream")},
    )

    # Import from schedule2 at DD (different values but no prior snapshot from schedule2)
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "DOOR NO.", "WIDTH", "HEIGHT", "FINISH", "MATERIAL", "HARDWARE SET", "FIRE RATING"
    ])
    ws.append([
        "Door 001", "3'-0\"", "7'-0\"", "stain", "aluminum", "HW-2", "60 min"
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    other_schedule_file = buf.getvalue()

    await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(other_schedule.id),
            "time_context_id": str(setup["dd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", other_schedule_file, "application/octet-stream")},
    )

    # Now import from schedule2 again at CD (same values as before in schedule2)
    resp = await client.post(
        "/api/v1/import",
        data={
            "source_item_id": str(other_schedule.id),
            "time_context_id": str(setup["cd_milestone"].id),
            "mapping_config": json.dumps(STANDARD_DOOR_MAPPING),
        },
        files={"file": ("schedule.xlsx", other_schedule_file, "application/octet-stream")},
    )
    assert resp.status_code == 201
    result = resp.json()

    # schedule2's values didn't change, so no changes detected (even though
    # they differ from schedule1)
    assert result["summary"]["source_changes"] == 0
    assert len(result["change_items"]) == 0
